from typing import Dict, Union, Type, List, Optional
from pydantic import BaseModel, validator, Extra, Field
import datetime
from openpyxl import load_workbook

from prodsim import adapters

from prodsim.data_structures import (
    queue_data,
    resource_data,
    time_model_data,
    state_data,
    processes_data,
    material_data,
    sink_data,
    source_data,
)

import pandas as pd


class SetupTime(BaseModel):
    machineSetupGroup: str
    taskSetupGroupFrom: str	
    taskSetupGroupTo: str
    Duration: datetime.datetime

    @validator("Duration", pre=True)
    def duration_to_timedelta(cls, v):
        return datetime.datetime.strptime(v[3:], "%H:%M:%S.%f")



class ProcessTime(BaseModel):
    machineDurationGroup: str
    taskDurationGroup: str
    duration: datetime.datetime

    @validator("duration", pre=True)
    def duration_to_timedelta(cls, v):
        return datetime.datetime.strptime(v[3:], "%H:%M:%S.%f")


class TransitionTime(BaseModel):
    jobTypeName: str
    transitionTimeGroup: str
    duration: datetime.datetime

    @validator("duration", pre=True)
    def duration_to_timedelta(cls, v: pd.Timestamp):
        if not v:
            return datetime.datetime.strptime("00:00:00.000", "%H:%M:%S.%f")
        return v


class TaskType(BaseModel):
    name: str
    label: Optional[str]
    description: str
    requiredCapabilityNames: str
    setupGroup: str
    durationGroup: str


class Machine(BaseModel):
    name: str
    label: str
    description: str
    locationName: str
    shiftName: str
    setupGroup: str
    availableCapabilityNames: str
    setupGroup: Optional[str]
    durationGroup: str

    class Config:
        extra = "ignore"


class JobWorkplan(BaseModel):
    jobTypeName: str
    taskTypeName: List[str]


class FlexisDataFrames(BaseModel):
    ApplicationName: pd.DataFrame
    Capability: pd.DataFrame
    Customer: pd.DataFrame
    Product: pd.DataFrame
    OrderType: pd.DataFrame
    JobType: pd.DataFrame
    Order: pd.DataFrame
    Location: pd.DataFrame
    Machine: pd.DataFrame
    TaskType: pd.DataFrame
    WorkPlan: pd.DataFrame
    SetupTime: pd.DataFrame
    ProcessTime: pd.DataFrame
    TransitionTime: pd.DataFrame
    CalendarParameter: pd.DataFrame
    DayPattern: pd.DataFrame
    Shift: pd.DataFrame
    NonworkingDay: pd.DataFrame
    Constraint: pd.DataFrame
    LinkType: pd.DataFrame
    Link: pd.DataFrame

    class Config:
        arbitrary_types_allowed = True


def read_excel(file_path: str):
    wb = load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames
    data = {}
    for sheet in sheets:
        data[sheet] = pd.DataFrame(
            wb[sheet].values,
            columns=[cell.value for cell in wb[sheet][1]],
        )
        data[sheet] = data[sheet].dropna(how="all")
        data[sheet] = data[sheet].iloc[1:]
        new_columns = [column for column in data[sheet].columns if column]
        data[sheet] = data[sheet][new_columns]
    return data


class FlexisAdapter(adapters.Adapter):
    _capability_process_dict: Dict[str, List[str]] = {}
    _flexis_data_frames: FlexisDataFrames = Field(default=None)

    class Config:
        extra = Extra.allow

    def read_data(self, file_path: str):
        input_data = read_excel(file_path)
        flexis_data_frames = FlexisDataFrames(**input_data)
        self._flexis_data_frames = flexis_data_frames
        self.initialize_time_models(flexis_data_frames)
        self.initialize_process_models(flexis_data_frames)
        self.initialize_resource_models(flexis_data_frames)
        self.initialize_setup_state_models(flexis_data_frames)
        self.initialize_queue_models()
        self.initialize_transport_models()
        self.initialize_material_models(flexis_data_frames)
        self.initialize_source_and_sink_data(flexis_data_frames)

    def get_object_from_data_frame(
        self, data_frame: pd.DataFrame, type: Type
    ) -> List[Type]:
        new_cols = [column for column in data_frame.columns if column]
        data_frame = data_frame[new_cols]
        data = data_frame.to_dict(orient="index")
        object_list = []
        for entry, entry_values in data.items():
            for key in list(entry_values):
                if ":" in key:
                    new_key = key.split(":")[0]
                    data[entry][new_key] = data[entry].pop(key)
            object_list.append(type(**data[entry]))

        return object_list

    def initialize_time_models(self, flexis_data_frames: FlexisDataFrames):
        process_time_data = self.get_object_from_data_frame(
            flexis_data_frames.ProcessTime, ProcessTime
        )
        for process_time in process_time_data:
            self.time_model_data.append(self.create_process_time_model(process_time))
        transition_time_data = self.get_object_from_data_frame(
            flexis_data_frames.TransitionTime, TransitionTime
        )
        self.time_model_data.append(
            self.create_transport_time_model()
        )

        setup_time_data = self.get_object_from_data_frame(
            flexis_data_frames.SetupTime, SetupTime
        )
        for setup_time in setup_time_data:
            self.time_model_data.append(
                self.create_setup_time_model(setup_time)
            )

    def create_process_time_model(self, process_time: ProcessTime):
        return time_model_data.FunctionTimeModelData(
            ID=process_time.machineDurationGroup + "-" + process_time.taskDurationGroup,
            description=f"Process time of {process_time.machineDurationGroup} and {process_time.taskDurationGroup}",
            type=time_model_data.TimeModelEnum.FunctionTimeModel,
            distribution_function="constant",
            parameters=[
                (
                    process_time.duration
                    - datetime.datetime.strptime("00:00:00.000", "%H:%M:%S.%f")
                ).total_seconds()
                / 60
            ],
            batch_size=100,
        )

    def create_transport_time_model(self):
        return time_model_data.ManhattanDistanceTimeModelData(
            ID="Transport",
            description=f"Transport time model",
            type=time_model_data.TimeModelEnum.ManhattanDistanceTimeModel,
            distribution_function="constant",
            speed=3*60,
            reaction_time=0.5
        )
    
    def create_setup_time_model(self, setup_time: SetupTime):
        return time_model_data.FunctionTimeModelData(
            ID=setup_time.machineSetupGroup + "-" + setup_time.taskSetupGroupFrom + "-" + setup_time.taskSetupGroupTo,
            description=f"Process time of {setup_time.machineSetupGroup} and {setup_time.taskSetupGroupFrom} and {setup_time.taskSetupGroupTo}",
            type=time_model_data.TimeModelEnum.FunctionTimeModel,
            distribution_function="constant",
            parameters=[
                (
                    setup_time.Duration
                    - datetime.datetime.strptime("00:00:00.000", "%H:%M:%S.%f")
                ).total_seconds()
                / 60
            ],
            batch_size=100,
        )


    def initialize_process_models(self, flexis_data_frames: FlexisDataFrames):
        task_type_data = self.get_object_from_data_frame(
            flexis_data_frames.TaskType, TaskType
        )
        machine_duration_groups = flexis_data_frames.Machine[
            "durationGroup:String"
        ].unique()
        for task_type in task_type_data:
            self.process_data += self.create_process_model(
                task_type, machine_duration_groups
            )

    def create_process_model(
        self, task_type: TaskType, machine_duration_groups: List[str]
    ):
        proceses_data_models = []
        time_model_ids = set([t.ID for t in self.time_model_data])
        for machine_duration_group in machine_duration_groups:
            time_model_id = (
                machine_duration_group.split(":")[0] + "-" + task_type.durationGroup
            )
            if time_model_id in time_model_ids:
                proceses_data_models.append(
                    processes_data.CapabilityProcessData(
                        ID=time_model_id,
                        description=task_type.description,
                        type=processes_data.ProcessTypeEnum.CapabilityProcesses,
                        time_model_id=time_model_id,
                        capability=task_type.requiredCapabilityNames,
                    )
                )
                if (
                    not task_type.requiredCapabilityNames
                    in self._capability_process_dict
                ):
                    self._capability_process_dict[
                        task_type.requiredCapabilityNames
                    ] = []
                self._capability_process_dict[task_type.requiredCapabilityNames].append(
                    time_model_id
                )
        return proceses_data_models
    

    def get_setup_for_processes(self, process1_repr, process2_repr, setup_time_data) -> SetupTime:
        for setup in setup_time_data:
            setup_from_variant = "_".join(setup.taskSetupGroupFrom.split("_")[1:])
            setup_to_variant = "_".join(setup.taskSetupGroupTo.split("_")[1:])
            if process1_repr == setup_from_variant and process2_repr == setup_to_variant:
                # print("Machted seutp", process1_repr, process2_repr)
                return setup
        for setup in setup_time_data:
            setup_to_variant = "_".join(setup.taskSetupGroupTo.split("_")[1:])
            if process2_repr == setup_to_variant and "*" == setup.taskSetupGroupFrom:
                # print("found generic from setup", process1_repr, process2_repr)
                return setup
        for setup in setup_time_data:
            if setup.taskSetupGroupTo == "*" and setup.taskSetupGroupFrom == "*":
                # print("found generic setup", process1_repr, process2_repr)
                return setup
    
    def initialize_setup_state_models(self, flexis_data_frames: FlexisDataFrames):
        setup_time_data = self.get_object_from_data_frame(
            flexis_data_frames.SetupTime, SetupTime
        )
        for resource in self.resource_data:
            for process1 in resource.processes:
                for process2 in resource.processes:
                    process1_repr = "_".join(process1.split("-")[1].split("_")[1:])
                    process2_repr = "_".join(process2.split("-")[1].split("_")[1:])
                    setup = self.get_setup_for_processes(process1_repr, process2_repr, setup_time_data)
                    setup_state = self.create_setup_state_model(process1, process2, setup)
                    self.state_data.append(setup_state)
                    resource.states.append(setup_state.ID)

    def create_setup_state_model(self, process1: str, process2: str, setup_time: SetupTime) -> state_data.SetupStateData:
        return state_data.SetupStateData(
            ID=setup_time.machineSetupGroup + "-" + setup_time.taskSetupGroupFrom + "-" + setup_time.taskSetupGroupTo,
            description=f"Process time of {setup_time.machineSetupGroup} and {setup_time.taskSetupGroupFrom} and {setup_time.taskSetupGroupTo}",
            time_model_id=setup_time.machineSetupGroup + "-" + setup_time.taskSetupGroupFrom + "-" + setup_time.taskSetupGroupTo,
            type=state_data.StateTypeEnum.SetupState,
            origin_setup=process1,
            target_setup=process2,
        )
    
    def initialize_resource_models(self, flexis_data_frames: FlexisDataFrames):
        resource_data = self.get_object_from_data_frame(
            flexis_data_frames.Machine, Machine
        )
        for resource in resource_data:
            self.resource_data.append(self.create_resource_model(resource))

    def create_resource_model(self, resource: Machine):
        return resource_data.ProductionResourceData(
            ID=resource.name,
            description=resource.description,
            capacity=1,
            location=(0, 0),
            controller="SimpleController",
            control_policy="FIFO",
            processes=self._capability_process_dict[resource.availableCapabilityNames],
            states=[],
        )

    def initialize_queue_models(self):
        for resource in self.resource_data:
            self.queue_data.append(
                queue_data.QueueData(
                    ID=resource.ID + "_input_queue",
                    description="Input queue for " + resource.ID,
                    capacity=0,
                )
            )
            self.queue_data.append(
                queue_data.QueueData(
                    ID=resource.ID + "_output_queue",
                    description="Output queue for " + resource.ID,
                    capacity=0,
                )
            )
            resource.input_queues = [resource.ID + "_input_queue"]
            resource.output_queues = [resource.ID + "_output_queue"]

    def initialize_transport_models(self):
        transport_process = processes_data.TransportProcessData(
            ID="TP1",
            description="Transport process",
            type=processes_data.ProcessTypeEnum.TransportProcesses,
            time_model_id="Transport",
        )
        self.process_data.append(transport_process)
        for i in range(2):
            transport_resource = resource_data.TransportResourceData(
                ID="Transport_" + str(i),
                description="Transport resource",
                capacity=1,
                location=(0, 0),
                controller="TransportController",
                control_policy="SPT_transport",
                processes=["TP1"],
                states=[],
            )
            self.resource_data.append(transport_resource)

    def get_capabilities_of_workplan(
        self, flexis_data_frames: FlexisDataFrames, workplan: JobWorkplan
    ):
        task_type_data = self.get_object_from_data_frame(
            flexis_data_frames.TaskType, TaskType
        )
        required_capability_names = []
        for task_type_name in workplan.taskTypeName:
            for task_type in task_type_data:
                if task_type_name == task_type.name:
                    possible_processes = self._capability_process_dict[task_type.requiredCapabilityNames]
                    possible_processes = [process for process in possible_processes if task_type_name in process]
                    required_capability_names.append(possible_processes[0])
                    break
            else:
                raise ValueError("Task type not found: " + task_type_name)
        return required_capability_names
    
    def initialize_material_models(self, flexis_data_frames: FlexisDataFrames):
        workplan_df = (
            flexis_data_frames.WorkPlan.groupby(by=["jobTypeName:String"])[
                "taskTypeName:String"
            ]
            .apply(list)
            .to_frame()
            .reset_index()
        )
        workplans: List[JobWorkplan] = self.get_object_from_data_frame(
            workplan_df, JobWorkplan
        )
        for workplan in workplans:
            required_capabilities = self.get_capabilities_of_workplan(
                flexis_data_frames, workplan
            )
            self.material_data.append(
                material_data.MaterialData(
                    ID=workplan.jobTypeName,
                    description=workplan.jobTypeName,
                    processes=required_capabilities,
                    transport_process="TP1",
                )
            )

    def initialize_source_and_sink_data(self, flexis_data_frames: FlexisDataFrames):
        # TODO: rework to scheduled state
        occurences = flexis_data_frames.Order["productName:String"].value_counts()
        occurences = occurences * 15
        freq = 48 * 60 / occurences
        for material in self.material_data:
            self.queue_data.append(
                queue_data.QueueData(
                    ID=material.ID + "_source_output_queue",
                    description="Source queue for " + material.ID,
                )
            )
            self.time_model_data.append(
                time_model_data.FunctionTimeModelData(
                    ID=material.ID + "_source_time_model",
                    description="Source time model for " + material.ID,
                    type=time_model_data.TimeModelEnum.FunctionTimeModel,
                    distribution_function="exponential",
                    parameters=[freq[material.ID]],
                )
            )

            self.source_data.append(
                source_data.SourceData(
                    ID=material.ID + "_source",
                    description="Source for " + material.ID,
                    capacity=100,
                    location=(0, 4.96),
                    material_type=material.ID,
                    time_model_id=material.ID + "_source_time_model",
                    router="CapabilityRouter",
                    # router="SimpleRouter", # also working
                    routing_heuristic="random",
                    output_queues=[material.ID + "_source_output_queue"],
                )
            )

            self.queue_data.append(
                queue_data.QueueData(
                    ID=material.ID + "_sink_input_queue",
                    description="Sink queue for " + material.ID,
                )
            )
            self.sink_data.append(
                sink_data.SinkData(
                    ID=material.ID + "_sink",
                    description="Sink for " + material.ID,
                    location=(0, 12.4),
                    material_type=material.ID,
                    input_queues=[
                        material.ID + "_sink_input_queue",
                    ],
                )
            )

    def write_data(self, file_path: str):
        new_machine_data = self._flexis_data_frames.Machine.copy()[0:0]
        for resource in self.resource_data:
            if isinstance(resource, resource_data.TransportResourceData):
                continue
            processes = [process for process in self.process_data if process.ID in resource.processes]
            capability = [process.capability for process in processes][0]
            row = self._flexis_data_frames.Machine.loc[self._flexis_data_frames.Machine["availableCapabilityNames:String"].str.contains(capability)][:1]
            row["name:String"] = resource.ID
            row["label:String"] = resource.ID
            row["location:Point"] = str(resource.location)
            new_machine_data = pd.concat([new_machine_data, row])
        new_machine_data.reset_index(inplace=True, drop=True)
        new_flexis_data_frames = self._flexis_data_frames.copy(deep=True)
        new_flexis_data_frames.Machine = new_machine_data
        writer = pd.ExcelWriter(file_path)
        values = new_flexis_data_frames.dict()
        for key, value in values.items():
            value.to_excel(writer, sheet_name=key, index=False)
        writer.close()
